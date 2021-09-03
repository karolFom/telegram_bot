import random
import telepot
import re
from collections import defaultdict
from openpyxl import load_workbook
from telepot.loop import MessageLoop
from telepot.namedtuple import InlineKeyboardButton, InlineKeyboardMarkup


class RandomBot:
    def __init__(self):
        self.token = '1674671265:AAGPxUBVD-efy_zuuDJXYS5kIM70EDhemvM'
        self.url = f'https://api.telegram.org/bot{self.token}/'
        self.bot = telepot.Bot(self.token)
        self.dict_kanto = defaultdict(dict)
        self.dict_oo = defaultdict(dict)
        self.reg = re.compile(r'\d*\s*\(.+\)')
        self.message_id = None

    def handle(self, msg):
        content_type, chat_type, chat_id = telepot.glance(msg)
        try:
            location = msg['text']
            if location == '/start_choosing':
                self.start_command(chat_id)
        except Exception:
            pass

    def start_command(self, chat_id):
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='Канто-Лес', callback_data='Канто-Лес'),
             InlineKeyboardButton(text='Канто-Поляна', callback_data='Канто-Поляна')],
            [InlineKeyboardButton(text='Канто-Пещеры', callback_data='Канто-Пещеры'),
             InlineKeyboardButton(text='Канто-Побережье', callback_data='Канто-Побережье')],
            [InlineKeyboardButton(text='Канто-Каменистая равнина', callback_data='Канто-Каменистая равнина'),
             InlineKeyboardButton(text='Канто-Океан', callback_data='Канто-Океан')],
            [InlineKeyboardButton(text='ОО-Дорога', callback_data='ОО-Дорога'),
             InlineKeyboardButton(text='ОО-Пустыня', callback_data='ОО-Пустыня')],
            [InlineKeyboardButton(text='ОО-Горная тропа', callback_data='ОО-Горная тропа'),
             InlineKeyboardButton(text='ОО-Лавандия', callback_data='ОО-Лавандия')],
            [InlineKeyboardButton(text='ОО-Лавовое плато', callback_data='ОО-Лавовое плато (Вулкан)'),
             InlineKeyboardButton(text='ОО-Грот', callback_data='ОО-Грот')],
            [InlineKeyboardButton(text='ОО-Горы', callback_data='ОО-Горы'),
             InlineKeyboardButton(text='ОО-Горное плато', callback_data='ОО-Горное плато')],
            [InlineKeyboardButton(text='ОО-Раскопки', callback_data='ОО-Раскопки'),
             InlineKeyboardButton(text='ОО-Степь', callback_data='ОО-Степь')],
            [InlineKeyboardButton(text='ОО-Пещера', callback_data='ОО-Пещера'),
             InlineKeyboardButton(text='ОО-Джунгли', callback_data='ОО-Джунгли')],
            [InlineKeyboardButton(text='ОО-Болото', callback_data='ОО-Болото'),
             InlineKeyboardButton(text='ОО-Затонувший корабль', callback_data='ОО-Затонувший корабль')],
            [InlineKeyboardButton(text='ОО-Ледник', callback_data='ОО-Ледник'),
             InlineKeyboardButton(text='ОО-Коралловый риф', callback_data='ОО-Коралловый риф')],
            [InlineKeyboardButton(text='ОО-Остров', callback_data='ОО-Остров'),
             InlineKeyboardButton(text='ОО-Океан', callback_data='ОО-Океан')],
            [InlineKeyboardButton(text='ОО-События для команды R', callback_data='ОО-События для команды R')],
        ]
        )
        response = self.bot.sendMessage(chat_id, 'Lets get started!', reply_markup=keyboard)
        self.message_id = response['message_id']

    def on_callback_query(self, msg):
        query_id, from_id, query_data = telepot.glance(msg, flavor='callback_query')
        self.bot.answerCallbackQuery(query_id)
        if self.message_id:
            self.bot.deleteMessage((from_id, self.message_id))
            self.message_id = None
        map, location = query_data.split('-')[0], query_data.split('-')[1]
        if map == 'ОО':
            events = self.dict_oo[location]
        else:
            events = self.dict_kanto[location]
        answer = random.choice(list(events.items()))
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text='Переиграть событие', callback_data=query_data)]
        ])
        if 'Разыгрываются из вкладки' in answer[1]:
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text='ОО-События для команды R', callback_data='ОО-События для команды R')]
            ])
        response = self.bot.sendMessage(from_id, f'Number: {answer[0]}\n {answer[1]}', reply_markup=keyboard)
        self.message_id = response['message_id']

    def loop(self):
        MessageLoop(self.bot, {'chat': self.handle,
                               'callback_query': self.on_callback_query}).run_as_thread()
        while True:
            n = input('To stop enter "stop":')
            if n.strip() == 'stop':
                break

    def parse_kanto(self):
        file = load_workbook('Kanto.xlsx', data_only=True)
        for sheet in file.worksheets:
            value = {}
            title = sheet.title
            for row in sheet.rows:
                vals = [cell.value for cell in row]
                value[vals[0]] = vals[2]
            new_value = {key: val for key, val in value.items() if isinstance(key, int) or re.match(self.reg, str(key))}
            self.dict_kanto[title] = new_value

    def parse_oo(self):
        file = load_workbook('OO.xlsx', data_only=True)
        for sheet in file.worksheets:
            value = {}
            title = sheet.title
            for row in sheet.iter_rows():
                vals = [cell.value for cell in row]
                value[vals[0]] = vals[2]
            new_value = {key: val for key, val in value.items() if isinstance(key, int) or re.match(self.reg, str(key))}
            self.dict_oo[title] = new_value


if __name__ == '__main__':
    bot = RandomBot()
    bot.parse_oo()
    bot.parse_kanto()
    bot.loop()







